%%writefile /content/rapport_journalier/sections/sva.py
import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    from utils.formatting import format_montant, format_quantite
except Exception:
    def format_montant(value):
        if value is None:
            return ""
        return f"{int(round(float(value))):,}".replace(",", " ") + " FCFA"

    def format_quantite(value):
        if value is None:
            return ""
        return f"{int(round(float(value))):,}".replace(",", " ")

DATA_DIR = Path("data")
DETAIL_FILE = DATA_DIR / "sva_detail.csv"

COMPONENTS = {
    "Djiguiya Data FS": "Djiguiya_Data_FS_TTC",
    "Djiguiya Voix FS": "Djiguiya_Voix_FS_TTC",
    "RBT": "RBT_CA",
    "Yotta": "Yotta_CA",
    "Megawin": "Megawin_CA",
    "ADN2": "ADN2_CA",
    "OFC": "OFC_CA",
    "Celebrite": "Celebrite_CA",
    "ODS": "ODS_CA",
    "CMS": "CMS_CA",
    "Youscribe": "Youscribe_CA",
    "Portail Islamique": "Portail_Islamique_CA",
    "Services Plus OMS": "Services_Plus_OMS_CA",
    "WIDO": "WIDO_CA",
    "E Carte": "E_Carte_CA",
    "Live Match": "Live_Match_CA",
    "Maxit TV": "Maxit_TV_CA",
    "Voxda": "Voxda_CA",
    "Mobibattle": "Mobibattle_CA",
    "FriendChat": "FriendChat_CA",
    "Autres SMS": "Autres_SMS_CA",
    "DV LIVE": "DV_LIVE_CA",
}

DETAIL_COLUMNS = [
    "Periode",
    "Djiguiya_Data_FS_TTC",
    "Djiguiya_Voix_FS_TTC",
    "RBT_Parc", "RBT_CA",
    "Yotta_Parc", "Yotta_CA",
    "Megawin_Parc", "Megawin_CA",
    "ADN2_Parc", "ADN2_CA",
    "OFC_Parc", "OFC_CA",
    "Celebrite_Parc", "Celebrite_CA",
    "ODS_Parc", "ODS_CA",
    "CMS_Parc", "CMS_CA",
    "Youscribe_Parc", "Youscribe_CA",
    "Portail_Islamique_Parc", "Portail_Islamique_CA",
    "Services_Plus_OMS_Parc", "Services_Plus_OMS_CA",
    "WIDO_Parc", "WIDO_CA",
    "E_Carte_Parc", "E_Carte_CA",
    "Live_Match_Parc", "Live_Match_CA",
    "Maxit_TV_Parc", "Maxit_TV_CA",
    "Voxda_Parc", "Voxda_CA",
    "Mobibattle_Parc", "Mobibattle_CA",
    "FriendChat_Parc", "FriendChat_CA",
    "Autres_SMS_Parc", "Autres_SMS_CA",
    "DV_LIVE_Parc", "DV_LIVE_CA",
    "SVA_Total_TTC",
    "SVA_Total_HT",
]

RENAME_MAP = {
    "Djiguiya_Data_FS_TTC": "Djiguiya Data FS TTC",
    "Djiguiya_Voix_FS_TTC": "Djiguiya Voix FS TTC",
    "RBT_Parc": "RBT Parc", "RBT_CA": "RBT CA",
    "Yotta_Parc": "Yotta Parc", "Yotta_CA": "Yotta CA",
    "Megawin_Parc": "Megawin Parc", "Megawin_CA": "Megawin CA",
    "ADN2_Parc": "ADN2 Parc", "ADN2_CA": "ADN2 CA",
    "OFC_Parc": "OFC Parc", "OFC_CA": "OFC CA",
    "Celebrite_Parc": "Celebrite Parc", "Celebrite_CA": "Celebrite CA",
    "ODS_Parc": "ODS Parc", "ODS_CA": "ODS CA",
    "CMS_Parc": "CMS Parc", "CMS_CA": "CMS CA",
    "Youscribe_Parc": "Youscribe Parc", "Youscribe_CA": "Youscribe CA",
    "Portail_Islamique_Parc": "Portail Islamique Parc", "Portail_Islamique_CA": "Portail Islamique CA",
    "Services_Plus_OMS_Parc": "Services Plus OMS Parc", "Services_Plus_OMS_CA": "Services Plus OMS CA",
    "WIDO_Parc": "WIDO Parc", "WIDO_CA": "WIDO CA",
    "E_Carte_Parc": "E Carte Parc", "E_Carte_CA": "E Carte CA",
    "Live_Match_Parc": "Live Match Parc", "Live_Match_CA": "Live Match CA",
    "Maxit_TV_Parc": "Maxit TV Parc", "Maxit_TV_CA": "Maxit TV CA",
    "Voxda_Parc": "Voxda Parc", "Voxda_CA": "Voxda CA",
    "Mobibattle_Parc": "Mobibattle Parc", "Mobibattle_CA": "Mobibattle CA",
    "FriendChat_Parc": "FriendChat Parc", "FriendChat_CA": "FriendChat CA",
    "Autres_SMS_Parc": "Autres SMS Parc", "Autres_SMS_CA": "Autres SMS CA",
    "DV_LIVE_Parc": "DV LIVE Parc", "DV_LIVE_CA": "DV LIVE CA",
    "SVA_Total_TTC": "SVA Total TTC",
    "SVA_Total_HT": "SVA Total HT",
}

def load_data():
    if not DETAIL_FILE.exists():
        raise FileNotFoundError(
            "Le fichier data/sva_detail.csv est introuvable. "
            "Crée-le avec la cellule de données fictives."
        )

    detail = pd.read_csv(DETAIL_FILE)
    detail["Date"] = pd.to_datetime(detail["Periode"], errors="coerce")
    detail = detail.dropna(subset=["Date"]).copy()

    required = ["Periode"] + list(COMPONENTS.values())
    missing = [col for col in required if col not in detail.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans sva_detail.csv : {missing}")

    numeric_cols = [c for c in detail.columns if c not in ["Periode", "Date"]]
    for col in numeric_cols:
        detail[col] = pd.to_numeric(detail[col], errors="coerce").fillna(0).astype(int)

    detail["SVA_Total_TTC"] = detail[list(COMPONENTS.values())].sum(axis=1).astype(int)
    detail["SVA_Total_HT"] = (detail["SVA_Total_TTC"] / 1.18).round().astype(int)

    summary = detail[["Periode", "Date", "SVA_Total_HT", "SVA_Total_TTC"]].copy()
    return summary, detail

def apply_periodicity(df, periodicite):
    out = df.copy()
    if periodicite == "Journalier":
        out["PeriodeAffichage"] = out["Date"].dt.strftime("%Y-%m-%d")
    else:
        out["PeriodeAffichage"] = out["Date"].dt.to_period("M").astype(str)
    return out

def aggregate_section(df, periodicite, cols):
    out = apply_periodicity(df, periodicite)
    return out.groupby("PeriodeAffichage", as_index=False)[cols].sum()

def to_long_ca(df, mapping, period_col="PeriodeAffichage"):
    rows = []
    for label, col in mapping.items():
        for _, row in df.iterrows():
            rows.append({
                "PeriodeAffichage": row[period_col],
                "Canal": label,
                "CA": row[col]
            })
    return pd.DataFrame(rows)

def get_last_30_days_table(df, end_date):
    end_date = pd.to_datetime(end_date)
    available = df[df["Date"] <= end_date].copy()

    if available.empty:
        return df.iloc[0:0].copy()

    start_date = max(available["Date"].min(), end_date - pd.Timedelta(days=29))
    return available[(available["Date"] >= start_date) & (available["Date"] <= end_date)].copy()

def render_html_table(df, group_styles, title):
    cols = list(df.columns)

    html = f'<div style="color:#FF7900;font-weight:700;font-size:22px;margin-bottom:8px;">{title}</div>'
    html += '<div style="overflow-x:auto;margin-bottom:24px;"><table style="border-collapse:collapse;width:100%;font-size:13px;">'
    html += '<thead><tr>'

    for col in cols:
        bg = "#dddddd"
        for prefix, color in group_styles.items():
            if col.startswith(prefix):
                bg = color
        html += f'<th style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{col}</th>'

    html += '</tr></thead><tbody>'

    for _, row in df.iterrows():
        html += '<tr>'
        for col in cols:
            bg = "#ffffff"
            for prefix, color in group_styles.items():
                if col.startswith(prefix):
                    bg = color
            val = row[col]
            if isinstance(val, (int, float)):
                val = format_quantite(val)
            html += f'<td style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{val}</td>'
        html += '</tr>'

    html += '</tbody></table></div>'
    st.markdown(html, unsafe_allow_html=True)

def hex_to_fill(hex_color):
    hex_color = hex_color.replace("#", "").upper()
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)

def style_sheet(ws, df, column_colors, title=None):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_offset = 1

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", start_color="FF7900", end_color="FF7900")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        row_offset = 2

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_offset, column=j, value=col)
        color = column_colors.get(col, "#DDDDDD")
        cell.fill = hex_to_fill(color)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, (_, row) in enumerate(df.iterrows(), start=row_offset + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=row[col])
            color = column_colors.get(col, "#FFFFFF")
            cell.fill = hex_to_fill(color)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for idx, col in enumerate(df.columns, start=1):
        letter = ws.cell(row=row_offset, column=idx).column_letter
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].tolist()))
        ws.column_dimensions[letter].width = min(max_len + 2, 24)

def build_excel_download(summary_df, detail_df, summary_colors, detail_colors):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Synthese"
    style_sheet(ws1, summary_df, summary_colors, title="SVA - SYNTHESE")

    ws2 = wb.create_sheet("Detail")
    style_sheet(ws2, detail_df, detail_colors, title="SVA - DETAIL")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def render():
    st.title("⭐ SVA")

    try:
        summary, detail = load_data()
    except Exception as e:
        st.error(f"Impossible de charger les données SVA : {e}")
        return

    if st.session_state.date_debut is None:
        st.session_state.date_debut = max(
            summary["Date"].max() - pd.Timedelta(days=29),
            summary["Date"].min()
        ).date()

    if st.session_state.date_fin is None:
        st.session_state.date_fin = summary["Date"].max().date()

    date_debut = pd.to_datetime(st.session_state.date_debut)
    date_fin = pd.to_datetime(st.session_state.date_fin)
    periodicite = st.session_state.periodicite

    summary_f = summary[(summary["Date"] >= date_debut) & (summary["Date"] <= date_fin)].copy()
    detail_f = detail[(detail["Date"] >= date_debut) & (detail["Date"] <= date_fin)].copy()

    if summary_f.empty:
        st.warning("Aucune donnée sur cette plage de dates.")
        return

    total_ht = int(summary_f["SVA_Total_HT"].sum())
    total_ttc = int(summary_f["SVA_Total_TTC"].sum())
    moyenne_journaliere = int(round(summary_f["SVA_Total_TTC"].mean()))

    c1, c2, c3 = st.columns(3)
    c1.metric("SVA Total HT", format_montant(total_ht))
    c2.metric("SVA Total TTC", format_montant(total_ttc))
    c3.metric("Moyenne journalière TTC", format_montant(moyenne_journaliere))

    st.markdown("---")

    with st.expander("🔽 Détail des composantes SVA", expanded=True):
        comp_totals = pd.DataFrame({
            "Type": list(COMPONENTS.keys()),
            "CA": [detail_f[col].sum() for col in COMPONENTS.values()]
        }).sort_values("CA", ascending=False)

        comp_totals = comp_totals[comp_totals["CA"] > 0].copy()

        top_n = 6
        top_components = comp_totals.head(top_n).copy()
        others_value = comp_totals.iloc[top_n:]["CA"].sum()

        pie_df = top_components.copy()
        if others_value > 0:
            pie_df = pd.concat(
                [pie_df, pd.DataFrame([{"Type": "Autres", "CA": others_value}])],
                ignore_index=True
            )

        top_labels = top_components["Type"].tolist()
        top_mapping = {label: COMPONENTS[label] for label in top_labels}

        detail_agg = aggregate_section(detail_f, periodicite, list(top_mapping.values()))
        detail_long = to_long_ca(detail_agg, top_mapping)

        a1, a2 = st.columns(2)

        with a1:
            fig1 = px.pie(
                pie_df,
                names="Type",
                values="CA",
                title="Répartition SVA (Top 6 + Autres)"
            )
            fig1.update_traces(textposition="inside", textinfo="percent+label")
            fig1.update_layout(template="plotly_dark", height=450)
            st.plotly_chart(fig1, use_container_width=True)

        with a2:
            fig2 = px.line(
                detail_long,
                x="PeriodeAffichage",
                y="CA",
                color="Canal",
                markers=True,
                title="Évolution des 6 composantes principales"
            )
            fig2.update_layout(template="plotly_dark", height=450)
            st.plotly_chart(fig2, use_container_width=True)

        st.markdown("#### Classement de toutes les composantes")
        fig3 = px.bar(
            comp_totals.sort_values("CA", ascending=True),
            x="CA",
            y="Type",
            orientation="h",
            title="CA par composante SVA"
        )
        fig3.update_layout(template="plotly_dark", height=650)
        st.plotly_chart(fig3, use_container_width=True)

    st.markdown("---")
    st.subheader("Tableaux de données")

    summary_table_f = get_last_30_days_table(summary, date_fin)
    detail_table_f = get_last_30_days_table(detail, date_fin)

    summary_display = summary_table_f.copy()
    summary_display["Periode"] = pd.to_datetime(summary_display["Periode"]).dt.strftime("%Y-%m-%d")
    summary_display = summary_display.rename(columns={
        "SVA_Total_HT": "SVA Total HT",
        "SVA_Total_TTC": "SVA Total TTC",
    })
    summary_display = summary_display[["Periode", "SVA Total HT", "SVA Total TTC"]]

    detail_display = detail_table_f[DETAIL_COLUMNS].copy()
    detail_display["Periode"] = pd.to_datetime(detail_display["Periode"]).dt.strftime("%Y-%m-%d")
    detail_display = detail_display.rename(columns=RENAME_MAP)

    summary_colors = {
        "Periode": "#F7F0D0",
        "SVA Total HT": "#D8E8BF",
        "SVA Total TTC": "#D7EBFB",
    }

    detail_colors = {
        "Periode": "#F7F0D0",
        "Djiguiya": "#F2DCCD",
        "RBT": "#D8E8BF",
        "Yotta": "#D7EBFB",
        "Megawin": "#EFC9AB",
        "ADN2": "#B7CCE3",
        "OFC": "#DCE8CC",
        "Celebrite": "#CBD3DF",
        "ODS": "#EFC7A6",
        "CMS": "#B8C9DE",
        "Youscribe": "#F4CCCC",
        "Portail": "#D9EAD3",
        "Services Plus": "#FFF2CC",
        "WIDO": "#D9D2E9",
        "E Carte": "#FCE5CD",
        "Live Match": "#CFE2F3",
        "Maxit": "#EAD1DC",
        "Voxda": "#D0E0E3",
        "Mobibattle": "#FFE599",
        "FriendChat": "#B6D7A8",
        "Autres SMS": "#C9DAF8",
        "DV LIVE": "#F9CB9C",
        "SVA Total": "#FFE6CC",
    }

    with st.expander("🔽 Tableau Synthèse SVA", expanded=False):
        render_html_table(summary_display, summary_colors, "SVA - SYNTHESE")

    with st.expander("🔽 Tableau Détail SVA", expanded=False):
        render_html_table(detail_display, detail_colors, "SVA - DETAIL")

    excel_bytes = build_excel_download(
        summary_df=summary_display,
        detail_df=detail_display,
        summary_colors=summary_colors,
        detail_colors=detail_colors,
    )

    st.download_button(
        label="📥 Télécharger les tableaux colorés (Excel)",
        data=excel_bytes,
        file_name="sva_tableaux_colores.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
