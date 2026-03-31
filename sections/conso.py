import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    from utils.formatting import format_montant, format_quantite
except Exception:
    def format_montant(value):
        if value is None:
            return ""
        return f"{int(round(float(value))):,}".replace(",", " ") + " FCFA"

    def format_quantite(value):
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:,.0f}".replace(",", " ")
        return f"{int(round(value)):,}".replace(",", " ")

DATA_DIR = Path("data")
DETAIL_FILE = DATA_DIR / "conso_detail.csv"

COMPONENTS = {
    "Conso TTC": "Conso_TTC",
    "Djiguiya TTC": "Djiguiya_TTC",
    "SEWA OMY TTC": "Sewa_OMY_TTC",
    "Data IM OMY TTC": "Data_IM_OMY_TTC",
    "Remb Djiguiya TTC": "Remb_Djiguiya_TTC",
    "Netaa Voix OMY TTC": "Netaa_Voix_OMY_TTC",
    "Netaa IM OMY TTC": "Netaa_IM_OMY_TTC",
    "Forfaits Voix Inter OMY TTC": "Forfaits_Voix_Inter_OMY_TTC",
    "Data Fixe OMY TTC": "Data_Fixe_OMY_TTC",
    "Topup SVA TTC": "Topup_SVA_TTC",
}


def load_data():
    if not DETAIL_FILE.exists():
        raise FileNotFoundError(
            "Le fichier data/conso_detail.csv est introuvable. "
            "Crée-le avec la cellule de données fictives."
        )

    detail = pd.read_csv(DETAIL_FILE)
    detail["Date"] = pd.to_datetime(detail["Periode"], errors="coerce")
    detail = detail.dropna(subset=["Date"]).copy()

    required = ["Periode"] + list(COMPONENTS.values())
    missing = [col for col in required if col not in detail.columns]
    if missing:
        raise ValueError(f"Colonnes manquantes dans conso_detail.csv : {missing}")

    for col in COMPONENTS.values():
        detail[col] = pd.to_numeric(detail[col], errors="coerce").fillna(0)

    detail["Total_Composantes_TTC"] = detail[list(COMPONENTS.values())].sum(axis=1)
    detail["Conso_Globale_HT"] = detail["Total_Composantes_TTC"] / 1.18

    summary = detail[["Periode", "Date", "Conso_Globale_HT", "Total_Composantes_TTC", "Conso_TTC"]].copy()
    return summary, detail


def apply_periodicity(df, periodicite):
    out = df.copy()
    if periodicite == "Journalier":
        out["PeriodeAffichage"] = out["Date"].dt.strftime("%Y-%m-%d")
    else:
        out["PeriodeAffichage"] = out["Date"].dt.to_period("M").astype(str)
    return out


def aggregate_section(df, periodicite, cols):
    out = apply_periodicity(df, periodicite)
    return out.groupby("PeriodeAffichage", as_index=False)[cols].sum()


def to_long_ca(df, mapping, period_col="PeriodeAffichage"):
    rows = []
    for label, col in mapping.items():
        for _, row in df.iterrows():
            rows.append({
                "PeriodeAffichage": row[period_col],
                "Canal": label,
                "CA": row[col]
            })
    return pd.DataFrame(rows)


def get_last_30_days_table(df, end_date):
    end_date = pd.to_datetime(end_date)
    available = df[df["Date"] <= end_date].copy()

    if available.empty:
        return df.iloc[0:0].copy()

    start_date = max(available["Date"].min(), end_date - pd.Timedelta(days=29))
    return available[(available["Date"] >= start_date) & (available["Date"] <= end_date)].copy()


def render_html_table(df, group_styles, title):
    cols = list(df.columns)

    html = f'<div style="color:#FF7900;font-weight:700;font-size:22px;margin-bottom:8px;">{title}</div>'
    html += '<div style="overflow-x:auto;margin-bottom:24px;"><table style="border-collapse:collapse;width:100%;font-size:13px;">'
    html += '<thead><tr>'

    for col in cols:
        bg = "#dddddd"
        for prefix, color in group_styles.items():
            if col.startswith(prefix):
                bg = color
        html += f'<th style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{col}</th>'

    html += '</tr></thead><tbody>'

    for _, row in df.iterrows():
        html += '<tr>'
        for col in cols:
            bg = "#ffffff"
            for prefix, color in group_styles.items():
                if col.startswith(prefix):
                    bg = color
            val = row[col]
            if isinstance(val, (int, float)):
                val = format_quantite(val)
            html += f'<td style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{val}</td>'
        html += '</tr>'

    html += '</tbody></table></div>'
    st.markdown(html, unsafe_allow_html=True)


def hex_to_fill(hex_color):
    hex_color = hex_color.replace("#", "").upper()
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)


def style_sheet(ws, df, column_colors, title=None):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_offset = 1

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", start_color="FF7900", end_color="FF7900")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        row_offset = 2

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_offset, column=j, value=col)
        color = column_colors.get(col, "#DDDDDD")
        cell.fill = hex_to_fill(color)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, (_, row) in enumerate(df.iterrows(), start=row_offset + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=row[col])
            color = column_colors.get(col, "#FFFFFF")
            cell.fill = hex_to_fill(color)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for idx, col in enumerate(df.columns, start=1):
        letter = ws.cell(row=row_offset, column=idx).column_letter
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].tolist()))
        ws.column_dimensions[letter].width = min(max_len + 2, 30)


def build_excel_download(summary_df, detail_df, summary_colors, detail_colors):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Synthese"
    style_sheet(ws1, summary_df, summary_colors, title="CONSO GLOBALE - SYNTHESE")

    ws2 = wb.create_sheet("Detail")
    style_sheet(ws2, detail_df, detail_colors, title="CONSO GLOBALE - DETAIL")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def render():
    st.title("📈 Conso")

    try:
        summary, detail = load_data()
    except Exception as e:
        st.error(f"Impossible de charger les données Conso : {e}")
        return

    if st.session_state.date_debut is None:
        st.session_state.date_debut = max(
            summary["Date"].max() - pd.Timedelta(days=29),
            summary["Date"].min()
        ).date()

    if st.session_state.date_fin is None:
        st.session_state.date_fin = summary["Date"].max().date()

    date_debut = pd.to_datetime(st.session_state.date_debut)
    date_fin = pd.to_datetime(st.session_state.date_fin)
    periodicite = st.session_state.periodicite

    summary_f = summary[(summary["Date"] >= date_debut) & (summary["Date"] <= date_fin)].copy()
    detail_f = detail[(detail["Date"] >= date_debut) & (detail["Date"] <= date_fin)].copy()

    if summary_f.empty:
        st.warning("Aucune donnée sur cette plage de dates.")
        return

    total_global_ht = summary_f["Conso_Globale_HT"].sum()
    total_ttc = summary_f["Total_Composantes_TTC"].sum()
    total_conso_ttc = summary_f["Conso_TTC"].sum()

    c1, c2, c3 = st.columns(3)
    c1.metric("CA Conso Globale HT", format_montant(total_global_ht))
    c2.metric("Total composantes TTC", format_montant(total_ttc))
    c3.metric("Conso TTC", format_montant(total_conso_ttc))

    st.markdown("---")

    with st.expander("🔽 Détail des composantes de la Conso globale", expanded=True):
        pie_df = pd.DataFrame({
            "Type": list(COMPONENTS.keys()),
            "CA": [detail_f[col].sum() for col in COMPONENTS.values()]
        })
        pie_df = pie_df[pie_df["CA"] > 0]

        detail_agg = aggregate_section(detail_f, periodicite, list(COMPONENTS.values()))
        detail_long = to_long_ca(detail_agg, COMPONENTS)

        a1, a2 = st.columns(2)

        with a1:
            fig3 = px.pie(
                pie_df,
                names="Type",
                values="CA",
                title="Poids des composantes TTC"
            )
            fig3.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig3, use_container_width=True)

        with a2:
            fig4 = px.line(
                detail_long,
                x="PeriodeAffichage",
                y="CA",
                color="Canal",
                markers=True,
                title="Évolution des composantes TTC"
            )
            fig4.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig4, use_container_width=True)

    st.markdown("---")
    st.subheader("Tableaux de données")

    summary_table_f = get_last_30_days_table(summary, date_fin)
    detail_table_f = get_last_30_days_table(detail, date_fin)

    summary_display = summary_table_f.copy()
    summary_display["Periode"] = pd.to_datetime(summary_display["Periode"]).dt.strftime("%Y-%m-%d")
    summary_display = summary_display.rename(columns={
        "Conso_Globale_HT": "Conso Globale HT",
        "Total_Composantes_TTC": "Total composantes TTC",
        "Conso_TTC": "Conso TTC",
    })
    summary_display = summary_display[["Periode", "Conso Globale HT", "Total composantes TTC", "Conso TTC"]]

    detail_display = detail_table_f.copy()
    detail_display["Periode"] = pd.to_datetime(detail_display["Periode"]).dt.strftime("%Y-%m-%d")
    keep_cols = ["Periode"] + list(COMPONENTS.values()) + ["Total_Composantes_TTC", "Conso_Globale_HT"]
    detail_display = detail_display[keep_cols].rename(columns={
        "Conso_TTC": "Conso TTC",
        "Djiguiya_TTC": "Djiguiya TTC",
        "Sewa_OMY_TTC": "SEWA OMY TTC",
        "Data_IM_OMY_TTC": "Data IM OMY TTC",
        "Remb_Djiguiya_TTC": "Remb Djiguiya TTC",
        "Netaa_Voix_OMY_TTC": "Netaa Voix OMY TTC",
        "Netaa_IM_OMY_TTC": "Netaa IM OMY TTC",
        "Forfaits_Voix_Inter_OMY_TTC": "Forfaits Voix Inter OMY TTC",
        "Data_Fixe_OMY_TTC": "Data Fixe OMY TTC",
        "Topup_SVA_TTC": "Topup SVA TTC",
        "Total_Composantes_TTC": "Total composantes TTC",
        "Conso_Globale_HT": "Conso Globale HT",
    })

    summary_colors = {
        "Periode": "#F7F0D0",
        "Conso Globale HT": "#D8E8BF",
        "Total composantes TTC": "#D7EBFB",
        "Conso TTC": "#F2DCCD",
    }

    detail_colors = {
        "Periode": "#F7F0D0",
        "Conso TTC": "#D8E8BF",
        "Djiguiya TTC": "#D7EBFB",
        "SEWA OMY TTC": "#F2DCCD",
        "Data IM OMY TTC": "#EFC9AB",
        "Remb Djiguiya TTC": "#B7CCE3",
        "Netaa Voix OMY TTC": "#DCE8CC",
        "Netaa IM OMY TTC": "#CBD3DF",
        "Forfaits Voix Inter OMY TTC": "#EFC7A6",
        "Data Fixe OMY TTC": "#B8C9DE",
        "Topup SVA TTC": "#B7B7B7",
        "Total composantes TTC": "#FFE6CC",
        "Conso Globale HT": "#FFE6CC",
    }

    with st.expander("🔽 Tableau Synthèse Conso", expanded=False):
        render_html_table(summary_display, summary_colors, "CONSO GLOBALE - SYNTHESE")

    with st.expander("🔽 Tableau Détail Conso", expanded=False):
        render_html_table(detail_display, detail_colors, "CONSO GLOBALE - DETAIL")

    excel_bytes = build_excel_download(
        summary_df=summary_display,
        detail_df=detail_display,
        summary_colors=summary_colors,
        detail_colors=detail_colors,
    )

    st.download_button(
        label="📥 Télécharger les tableaux colorés (Excel)",
        data=excel_bytes,
        file_name="conso_tableaux_colores.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
