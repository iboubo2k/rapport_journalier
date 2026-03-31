import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from utils.formatting import format_montant, format_quantite

DATA_DIR = Path("data")


def load_data():
    summary = pd.read_csv(DATA_DIR / "recharge_summary.csv")
    direct = pd.read_csv(DATA_DIR / "recharge_direct.csv")
    indirect = pd.read_csv(DATA_DIR / "recharge_indirect.csv")

    for df in [summary, direct, indirect]:
        df["Date"] = pd.to_datetime(df["Periode"], errors="coerce")

    return summary, direct, indirect


def apply_periodicity(df, periodicite):
    out = df.copy()
    if periodicite == "Journalier":
        out["PeriodeAffichage"] = out["Date"].dt.strftime("%Y-%m-%d")
    else:
        out["PeriodeAffichage"] = out["Date"].dt.to_period("M").astype(str)
    return out


def aggregate_summary(summary_df, periodicite):
    df = apply_periodicity(summary_df, periodicite)
    return df.groupby("PeriodeAffichage", as_index=False)[["Directes", "Indirectes"]].sum()


def aggregate_section(section_df, periodicite, channel_cols):
    df = apply_periodicity(section_df, periodicite)
    return df.groupby("PeriodeAffichage", as_index=False)[channel_cols].sum()


def to_long_ca(df, mapping, period_col="PeriodeAffichage"):
    rows = []
    for label, ca_col in mapping.items():
        for _, row in df.iterrows():
            rows.append({
                "PeriodeAffichage": row[period_col],
                "Canal": label,
                "CA": row[ca_col]
            })
    return pd.DataFrame(rows)


def get_last_30_days_table(df, end_date):
    end_date = pd.to_datetime(end_date)

    available = df[df["Date"] <= end_date].copy()

    if available.empty:
        return df.iloc[0:0].copy()

    start_date = max(
        available["Date"].min(),
        end_date - pd.Timedelta(days=29)
    )

    return available[
        (available["Date"] >= start_date) & (available["Date"] <= end_date)
    ].copy()


def render_html_table(df, group_styles, title):
    cols = list(df.columns)

    html = f'<div style="color:#FF7900;font-weight:700;font-size:22px;margin-bottom:8px;">{title}</div>'
    html += '<div style="overflow-x:auto;margin-bottom:24px;"><table style="border-collapse:collapse;width:100%;font-size:13px;">'
    html += '<thead><tr>'

    for col in cols:
        bg = "#dddddd"
        for prefix, color in group_styles.items():
            if col.startswith(prefix):
                bg = color
        html += f'<th style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{col}</th>'

    html += '</tr></thead><tbody>'

    for _, row in df.iterrows():
        html += '<tr>'
        for col in cols:
            bg = "#ffffff"
            for prefix, color in group_styles.items():
                if col.startswith(prefix):
                    bg = color
            val = row[col]
            if isinstance(val, (int, float)):
                val = format_quantite(val)
            html += f'<td style="background:{bg};border:1px solid #333;padding:6px 8px;text-align:center;color:black;">{val}</td>'
        html += '</tr>'

    html += '</tbody></table></div>'
    st.markdown(html, unsafe_allow_html=True)


def hex_to_fill(hex_color):
    hex_color = hex_color.replace("#", "").upper()
    return PatternFill(fill_type="solid", start_color=hex_color, end_color=hex_color)


def style_sheet(ws, df, column_colors, title=None):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row_offset = 1

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", start_color="FF7900", end_color="FF7900")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        row_offset = 2

    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_offset, column=j, value=col)
        color = column_colors.get(col, "#DDDDDD")
        cell.fill = hex_to_fill(color)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, (_, row) in enumerate(df.iterrows(), start=row_offset + 1):
        for j, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=i, column=j, value=row[col])
            color = column_colors.get(col, "#FFFFFF")
            cell.fill = hex_to_fill(color)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for idx, col in enumerate(df.columns, start=1):
        letter = ws.cell(row=row_offset, column=idx).column_letter
        max_len = max(len(str(col)), *(len(str(v)) for v in df[col].tolist()))
        ws.column_dimensions[letter].width = min(max_len + 2, 24)


def build_excel_download(summary_df, direct_df, indirect_df, summary_colors, direct_colors, indirect_colors):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Recap"
    style_sheet(ws1, summary_df, summary_colors, title="RECHARGES CA TTC")

    ws2 = wb.create_sheet("Directes")
    style_sheet(ws2, direct_df, direct_colors, title="Recharges Directes")

    ws3 = wb.create_sheet("Indirectes")
    style_sheet(ws3, indirect_df, indirect_colors, title="Recharges Indirectes")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def render():
    st.title("💳 Recharge")

    summary, direct, indirect = load_data()

    if st.session_state.date_debut is None:
        st.session_state.date_debut = max(
            summary["Date"].max() - pd.Timedelta(days=29),
            summary["Date"].min()
        ).date()

    if st.session_state.date_fin is None:
        st.session_state.date_fin = summary["Date"].max().date()

    date_debut = pd.to_datetime(st.session_state.date_debut)
    date_fin = pd.to_datetime(st.session_state.date_fin)
    periodicite = st.session_state.periodicite

    # Filtres pour KPI + graphes
    summary_f = summary[(summary["Date"] >= date_debut) & (summary["Date"] <= date_fin)].copy()
    direct_f = direct[(direct["Date"] >= date_debut) & (direct["Date"] <= date_fin)].copy()
    indirect_f = indirect[(indirect["Date"] >= date_debut) & (indirect["Date"] <= date_fin)].copy()

    if summary_f.empty:
        st.warning("Aucune donnée sur cette plage de dates.")
        return

    total_direct = summary_f["Directes"].sum()
    total_indirect = summary_f["Indirectes"].sum()
    total_all = total_direct + total_indirect

    c1, c2, c3 = st.columns(3)
    c1.metric("CA Recharges Directes", format_montant(total_direct))
    c2.metric("CA Recharges Indirectes", format_montant(total_indirect))
    c3.metric("CA Recharge Total", format_montant(total_all))

    st.markdown("---")
    st.subheader(f"Vue globale Recharge — {periodicite}")

    summary_agg = aggregate_summary(summary_f, periodicite)

    g1, g2 = st.columns(2)

    with g1:
        camembert_df = pd.DataFrame({
            "Type": ["Directes", "Indirectes"],
            "CA": [total_direct, total_indirect]
        })
        fig = px.pie(
            camembert_df,
            names="Type",
            values="CA",
            title="Répartition Directes vs Indirectes"
        )
        fig.update_layout(template="plotly_dark", height=420)
        st.plotly_chart(fig, use_container_width=True)

    with g2:
        evol_df = summary_agg.melt(
            id_vars="PeriodeAffichage",
            value_vars=["Directes", "Indirectes"],
            var_name="Type",
            value_name="CA"
        )
        fig2 = px.line(
            evol_df,
            x="PeriodeAffichage",
            y="CA",
            color="Type",
            markers=True,
            title="Évolution des Recharges Directes et Indirectes"
        )
        fig2.update_layout(template="plotly_dark", height=420)
        st.plotly_chart(fig2, use_container_width=True)

    with st.expander("🔽 Détail Recharges Directes", expanded=True):
        direct_map = {
            "CAG": "Ca_CAG",
            "NAFAMA": "Ca_NAFAMA",
            "OMY HORS APPLI": "Ca_OMY_HORS",
            "OMY APPLI": "Ca_OMY_APPLI",
            "OMY APPLI WAVE": "Ca_OMY_WAVE",
            "Transfert Pays": "Ca_TRANSFERT_PAYS",
            "Single Wallet": "Ca_SINGLE_WALLET",
        }

        direct_camembert = pd.DataFrame({
            "Canal": list(direct_map.keys()),
            "CA": [direct_f[col].sum() for col in direct_map.values()]
        })

        direct_agg = aggregate_section(direct_f, periodicite, list(direct_map.values()))
        direct_long = to_long_ca(direct_agg, direct_map)

        a1, a2 = st.columns(2)

        with a1:
            fig3 = px.pie(
                direct_camembert,
                names="Canal",
                values="CA",
                title="Répartition des Recharges Directes"
            )
            fig3.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig3, use_container_width=True)

        with a2:
            fig4 = px.line(
                direct_long,
                x="PeriodeAffichage",
                y="CA",
                color="Canal",
                markers=True,
                title="Évolution des Recharges Directes"
            )
            fig4.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig4, use_container_width=True)

    with st.expander("🔽 Détail Recharges Indirectes", expanded=True):
        indirect_map = {
            "IM": "Ca_IM",
            "SEVA": "Ca_SEVA",
            "NETAA": "Ca_NETAA",
            "BOX": "Ca_BOX",
            "Forfaits VOIX": "Ca_FORFAITS_VOIX",
            "Services Plus": "Ca_SERVICES_PLUS",
            "Fibre Optique": "Ca_FIBRE_OPTIQUE",
        }

        indirect_camembert = pd.DataFrame({
            "Canal": list(indirect_map.keys()),
            "CA": [indirect_f[col].sum() for col in indirect_map.values()]
        })

        indirect_agg = aggregate_section(indirect_f, periodicite, list(indirect_map.values()))
        indirect_long = to_long_ca(indirect_agg, indirect_map)

        b1, b2 = st.columns(2)

        with b1:
            fig5 = px.pie(
                indirect_camembert,
                names="Canal",
                values="CA",
                title="Répartition des Recharges Indirectes"
            )
            fig5.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig5, use_container_width=True)

        with b2:
            fig6 = px.line(
                indirect_long,
                x="PeriodeAffichage",
                y="CA",
                color="Canal",
                markers=True,
                title="Évolution des Recharges Indirectes"
            )
            fig6.update_layout(template="plotly_dark", height=420)
            st.plotly_chart(fig6, use_container_width=True)

    st.markdown("---")
    st.subheader("Tableaux de données")

    summary_colors = {
        "Periode": "#F7F0D0", "Date": "#F7F0D0",
        "Directes": "#D8E8BF", "Indirectes": "#D7EBFB"
    }

    direct_colors = {
        "Periode": "#F7F0D0", "Date": "#F7F0D0",
        "Parc_CAG": "#D9D9D9", "Ca_CAG": "#D9D9D9",
        "Parc_NAFAMA": "#B7CCE3", "Ca_NAFAMA": "#B7CCE3",
        "Parc_OMY_HORS": "#F2DCCD", "Ca_OMY_HORS": "#F2DCCD",
        "Parc_OMY_APPLI": "#9FBBE7", "Ca_OMY_APPLI": "#9FBBE7",
        "Parc_OMY_WAVE": "#CBD3DF", "Ca_OMY_WAVE": "#CBD3DF",
        "Parc_TRANSFERT_PAYS": "#EFC9AB", "Ca_TRANSFERT_PAYS": "#EFC9AB",
        "Parc_SINGLE_WALLET": "#DCE8CC", "Ca_SINGLE_WALLET": "#DCE8CC",
    }

    indirect_colors = {
        "Periode": "#F7F0D0", "Date": "#F7F0D0",
        "Parc_IM": "#B8C9DE", "Ca_IM": "#B8C9DE",
        "Parc_SEVA": "#EFC7A6", "Ca_SEVA": "#EFC7A6",
        "Parc_NETAA": "#B7B7B7", "Ca_NETAA": "#B7B7B7",
        "Parc_BOX": "#9EB38F", "Ca_BOX": "#9EB38F",
        "Parc_FORFAITS_VOIX": "#CBD3DF", "Ca_FORFAITS_VOIX": "#CBD3DF",
        "Parc_SERVICES_PLUS": "#EFC7A6", "Ca_SERVICES_PLUS": "#EFC7A6",
        "Parc_FIBRE_OPTIQUE": "#DCE8CC", "Ca_FIBRE_OPTIQUE": "#DCE8CC",
    }

    # Filtres spéciaux pour les tableaux du bas :
    # - par défaut : 30 derniers jours disponibles
    # - si période choisie : 30 derniers jours jusqu'à date_fin
    summary_table_f = get_last_30_days_table(summary, date_fin)
    direct_table_f = get_last_30_days_table(direct, date_fin)
    indirect_table_f = get_last_30_days_table(indirect, date_fin)

    summary_display = summary_table_f.copy()
    summary_display["Date"] = summary_display["Date"].dt.strftime("%Y-%m-%d")
    summary_display["Directes"] = summary_display["Directes"].apply(format_quantite)
    summary_display["Indirectes"] = summary_display["Indirectes"].apply(format_quantite)

    direct_display = direct_table_f.copy()
    direct_display["Date"] = direct_display["Date"].dt.strftime("%Y-%m-%d")

    indirect_display = indirect_table_f.copy()
    indirect_display["Date"] = indirect_display["Date"].dt.strftime("%Y-%m-%d")

    with st.expander("🔽 Tableau RECHARGES CA TTC", expanded=False):
        render_html_table(summary_display, summary_colors, "RECHARGES CA TTC")

    with st.expander("🔽 Tableau Recharges Directes", expanded=False):
        render_html_table(direct_display, direct_colors, "Recharges Directes")

    with st.expander("🔽 Tableau Recharges Indirectes", expanded=False):
        render_html_table(indirect_display, indirect_colors, "Recharges Indirectes")

    excel_bytes = build_excel_download(
        summary_df=summary_display,
        direct_df=direct_display,
        indirect_df=indirect_display,
        summary_colors=summary_colors,
        direct_colors=direct_colors,
        indirect_colors=indirect_colors
    )

    st.download_button(
        label="📥 Télécharger les tableaux colorés (Excel)",
        data=excel_bytes,
        file_name="recharges_tableaux_colores.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
